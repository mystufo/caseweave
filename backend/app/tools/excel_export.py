"""Export test cases to Excel (.xlsx)."""
import io
from typing import Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADERS = [
    "用例编号", "用例名称", "用例模块", "优先级", "前置条件",
    "执行步骤", "预期结果", "备注", "自测结果",
]

FIELD_MAP = [
    "case_number", "name", "module", "priority", "preconditions",
    "steps", "expected_result", "remarks", "test_result",
]

COLUMN_WIDTHS = [15, 35, 15, 8, 30, 50, 40, 25, 12]


def _style_header_row(ws, row: int) -> None:
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def _write_cases_to_sheet(ws, cases: list[dict[str, Any]]) -> None:
    _style_header_row(ws, 1)

    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_idx, case in enumerate(cases, start=2):
        for col_idx, field in enumerate(FIELD_MAP, start=1):
            value = case.get(field) or ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    # Set column widths
    for col_idx, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Set row height for data rows
    for row_idx in range(2, len(cases) + 2):
        ws.row_dimensions[row_idx].height = 45


def export_test_cases(
    cases: list[dict[str, Any]],
    group_by_module: bool = True,
) -> bytes:
    """
    Export test cases to Excel bytes.

    If group_by_module=True, each module gets its own sheet.
    Returns raw bytes of the .xlsx file.
    """
    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    if group_by_module:
        # Group by module
        from collections import defaultdict
        grouped: dict[str, list] = defaultdict(list)
        for case in cases:
            module = case.get("module") or "其他"
            grouped[module].append(case)

        for module_name, module_cases in grouped.items():
            sheet_name = module_name[:31]  # Excel sheet name limit
            ws = wb.create_sheet(title=sheet_name)
            _write_cases_to_sheet(ws, module_cases)
    else:
        ws = wb.create_sheet(title="测试用例")
        _write_cases_to_sheet(ws, cases)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
